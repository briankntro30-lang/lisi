import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle, Patch
from datetime import datetime
import sqlite3
import json
import io
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Simogramme", layout="wide")

st.markdown("""
<style>
.main { background-color: #f4f6f9; }
h1, h2, h3 { color: #1f2937; font-weight: 700; }
.stButton>button {
    background-color: #1f2937; color: white;
    border-radius: 8px; height: 45px; font-weight: bold; border: none;
}
.stButton>button:hover { background-color: #374151; }
.metric-card {
    background-color: white; padding: 15px; border-radius: 10px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1); text-align: center;
}
.metric-value { font-size: 32px; font-weight: bold; color: #1f2937; }
.metric-label { font-size: 14px; color: #6b7280; margin-top: 5px; }
.metric-delta { font-size: 12px; margin-top: 5px; }
.metric-card-repos {
    background-color: #fff7ed; padding: 15px; border-radius: 10px;
    box-shadow: 0 1px 3px rgba(249,115,22,0.2); text-align: center;
    border: 2px solid #f97316;
}
.metric-value-repos { font-size: 32px; font-weight: bold; color: #f97316; }
.metric-label-repos { font-size: 14px; color: #92400e; margin-top: 5px; }
.metric-delta-repos { font-size: 12px; margin-top: 5px; color: #c2410c; }
.metric-card-m1 {
    background-color: #eff6ff; padding: 15px; border-radius: 10px;
    box-shadow: 0 1px 3px rgba(59,130,246,0.2); text-align: center;
    border: 2px solid #3b82f6;
}
.metric-value-m1 { font-size: 28px; font-weight: bold; color: #1d4ed8; }
.metric-label-m1 { font-size: 13px; color: #1e40af; margin-top: 5px; }
.metric-delta-m1 { font-size: 11px; margin-top: 5px; color: #3b82f6; }
.metric-card-m2 {
    background-color: #f0fdf4; padding: 15px; border-radius: 10px;
    box-shadow: 0 1px 3px rgba(34,197,94,0.2); text-align: center;
    border: 2px solid #22c55e;
}
.metric-value-m2 { font-size: 28px; font-weight: bold; color: #15803d; }
.metric-label-m2 { font-size: 13px; color: #166534; margin-top: 5px; }
.metric-delta-m2 { font-size: 11px; margin-top: 5px; color: #22c55e; }
.info-icon {
    display: inline-block; width: 16px; height: 16px;
    background-color: #6b7280; color: white; border-radius: 50%;
    text-align: center; font-size: 11px; font-weight: bold;
    line-height: 16px; margin-left: 5px; cursor: help; font-family: monospace;
}
.legend-color {
    display: inline-block; width: 20px; height: 20px;
    border-radius: 3px; margin-right: 5px; vertical-align: middle;
}
.sim-card {
    background: white; border-radius: 14px; padding: 18px 22px;
    margin-bottom: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    border-left: 6px solid #f97316; transition: box-shadow 0.2s;
}
.sim-card:hover { box-shadow: 0 4px 18px rgba(0,0,0,0.13); }
.sim-article { font-size: 28px; font-weight: 800; color: #f97316; letter-spacing: 1px; margin-bottom: 2px; }
.sim-date { font-size: 11px; color: #9ca3af; margin-bottom: 10px; }
.sim-row { display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 6px; }
.sim-badge { background: #f3f4f6; border-radius: 6px; padding: 3px 10px; font-size: 12px; color: #374151; }
.sim-badge b { color: #1f2937; }
.sim-kpi { display: flex; gap: 18px; margin-top: 10px; padding-top: 10px; border-top: 1px solid #f3f4f6; }
.sim-kpi-item { text-align: center; }
.sim-kpi-val { font-size: 18px; font-weight: 700; color: #1f2937; }
.sim-kpi-lbl { font-size: 10px; color: #9ca3af; }
.chrono-header {
    background: linear-gradient(90deg,#1f2937,#374151);
    color: white; border-radius: 10px; padding: 12px 18px; margin-bottom: 12px;
    font-weight: 700; font-size: 16px;
}
.chrono-stats {
    background: #f0fdf4; border: 1px solid #bbf7d0;
    border-radius: 8px; padding: 10px 14px; font-size: 13px; color: #166534;
}
.repos-banner {
    background: linear-gradient(135deg, #fff7ed 0%, #ffedd5 100%);
    border: 2px solid #f97316; border-radius: 14px; padding: 18px 24px;
    margin: 16px 0; display: flex; align-items: center; gap: 24px;
}
.repos-icon { font-size: 42px; }
.repos-title { font-size: 13px; color: #92400e; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; }
.repos-val { font-size: 36px; font-weight: 800; color: #f97316; }
.repos-sub { font-size: 12px; color: #c2410c; margin-top: 2px; }
/* Chrono machine badge */
.mach-badge-op  { background:#e0e7ff; color:#3730a3; font-weight:700; border-radius:5px; padding:2px 8px; font-size:11px; }
.mach-badge-m1  { background:#dbeafe; color:#1d4ed8; font-weight:700; border-radius:5px; padding:2px 8px; font-size:11px; }
.mach-badge-m2  { background:#dcfce7; color:#15803d; font-weight:700; border-radius:5px; padding:2px 8px; font-size:11px; }
.cycle-header-m1 {
    background: linear-gradient(90deg,#1d4ed8,#3b82f6);
    color:white; border-radius:10px; padding:10px 16px;
    font-weight:700; font-size:15px; margin: 12px 0 8px 0;
}
.cycle-header-m2 {
    background: linear-gradient(90deg,#15803d,#22c55e);
    color:white; border-radius:10px; padding:10px 16px;
    font-weight:700; font-size:15px; margin: 12px 0 8px 0;
}
</style>
""", unsafe_allow_html=True)

LOGO_URL = "https://th.bing.com/th/id/R.0a38b5bebde3a9c6b070c0ad42c162d3?rik=U63XkDE5XvdVCg&riu=http%3a%2f%2fbandemfg.com%2fimages%2ffooter-logo.png&ehk=NquqcRNMxNTQUwJ5DrA7Sz1HroAbEmUUL7LemhCeyCQ%3d&risl=&pid=ImgRaw&r=0"
st.image(LOGO_URL, width=250)

# ===================================================
# DATABASE
# ===================================================

DB_PATH = r"C:\Users\BFRANCOCANTERO\Downloads\Data\simogramme_data.db"
try:
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    open(DB_PATH, 'a').close()
except Exception:
    DB_PATH = os.path.join(os.path.expanduser("~"), "simogramme_data.db")

def init_database():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS configurations
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  date TEXT, numero_of TEXT, reference_piece TEXT, numero_machine TEXT,
                  pdc TEXT, numero_article TEXT,
                  coef_temps_humain REAL, coef_temps_cycle REAL,
                  heures_travail REAL, machines TEXT, donnees TEXT, resultats TEXT)''')
    cols = [r[1] for r in conn.execute("PRAGMA table_info(configurations)").fetchall()]
    for col, typ in [("numero_of","TEXT"), ("numero_article","TEXT"),
                     ("coef_temps_humain","REAL"), ("coef_temps_cycle","REAL")]:
        if col not in cols:
            conn.execute(f"ALTER TABLE configurations ADD COLUMN {col} {typ} DEFAULT 1.0")
    conn.commit()
    conn.close()

def save_configuration(data):
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute("""INSERT INTO configurations
                     (date, numero_of, reference_piece, numero_machine, pdc, numero_article,
                      coef_temps_humain, coef_temps_cycle, heures_travail, machines, donnees, resultats)
                     VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                  (data['date'], data['numero_of'], data['reference_piece'],
                   data['numero_machine'], data['pdc'], data['numero_article'],
                   data['coef_temps_humain'], data['coef_temps_cycle'],
                   data['heures_travail'], data['machines'], data['donnees'], data['resultats']))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Erreur sauvegarde: {e}")
        return False

def load_configurations():
    try:
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql_query("SELECT * FROM configurations ORDER BY date DESC", conn)
        conn.close()
        return df
    except Exception as e:
        st.error(f"Erreur chargement: {e}")
        return pd.DataFrame()

def delete_configuration(config_id):
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute('DELETE FROM configurations WHERE id = ?', (config_id,))
        conn.commit()
        conn.close()
    except Exception as e:
        st.error(f"Erreur suppression: {e}")

init_database()

# ===================================================
# EXCEL EXPORT
# ===================================================

def build_excel(edited_df, machines, sidebar_info, resultats, img_bytes):
    wb = Workbook()
    DARK="1F2937"; ACCENT="F97316"; LIGHT="F3F4F6"; WHITE="FFFFFF"

    def hdr(ws, row, col, value, bg=DARK, fg=WHITE, bold=True, size=11, align="center", border=True):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if border:
            thin = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return cell

    def val(ws, row, col, value, bg=WHITE, bold=False, align="center", color="1F2937"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", bold=bold, color=color, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return cell

    ws = wb.active; ws.title = "Synthèse"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.merge_cells("A2:B2")
    hdr(ws,2,1,"RAPPORT SIMOGRAMME",bg=DARK,fg=WHITE,bold=True,size=14,align="center",border=False)
    ws.row_dimensions[2].height = 36
    ws.merge_cells("A3:B3")
    c = ws.cell(row=3,column=1,value=f"Article: {sidebar_info.get('numero_article','')}   |   OF: {sidebar_info['numero_of']}   |   Machine: {sidebar_info['numero_machine']}")
    c.font=Font(name="Arial",color="6B7280",size=10)
    c.fill=PatternFill("solid",fgColor=LIGHT)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[3].height=20; ws.row_dimensions[4].height=8
    hdr(ws,5,1,"IDENTIFICATION",bg=ACCENT); hdr(ws,5,2,"Valeur",bg=ACCENT)
    ident = [("Numéro d'article",sidebar_info.get("numero_article","")),
             ("Numéro OF",sidebar_info["numero_of"]),
             ("Référence pièce",sidebar_info["reference_piece"]),
             ("Numéro machine",sidebar_info["numero_machine"]),
             ("PDC",sidebar_info["pdc"]),
             ("Date",sidebar_info["date"])]
    for i,(k,v) in enumerate(ident):
        bg=WHITE if i%2==0 else LIGHT
        val(ws,6+i,1,k,bg=bg,align="left"); val(ws,6+i,2,v,bg=bg)
    r=6+len(ident)+1; ws.row_dimensions[r-1].height=8
    hdr(ws,r,1,"COEFFICIENTS",bg=ACCENT); hdr(ws,r,2,"Valeur",bg=ACCENT)
    coefs=[("Coef. Temps Humain",sidebar_info["coef_temps_humain"]),
           ("Coef. Temps Cycle",sidebar_info["coef_temps_cycle"]),
           ("H/jour",sidebar_info["heures_travail"])]
    for i,(k,v) in enumerate(coefs):
        bg=WHITE if i%2==0 else LIGHT
        val(ws,r+1+i,1,k,bg=bg,align="left"); val(ws,r+1+i,2,v,bg=bg)
    r2=r+1+len(coefs)+1; ws.row_dimensions[r2-1].height=8
    hdr(ws,r2,1,"RÉSULTATS",bg=ACCENT); hdr(ws,r2,2,"Valeur",bg=ACCENT)
    res=resultats
    # Use M1 results for main sheet, M2 appended below if present
    results=[("Temps cycle final (s)",round(res.get("temps_cycle_final",0),4)),
             ("Temps cycle final (UM)",round(res.get("temps_cycle_final",0)/36,4)),
             ("Temps machine (s)",round(res.get("total_machine_time",0),4)),
             ("Temps manuel TM (s)",round(res.get("total_operator_manual",0),4)),
             ("Temps parallèle TTM (s)",round(res.get("total_operator_parallel",0),4)),
             ("Temps repos TR (s)",round(res.get("total_repos_time",0),4)),
             ("Temps masqué TZ (s)",round(res.get("total_masked_time",0),4)),
             ("Taux occ. opérateur %",round(res.get("taux_h",0),2)),
             ("Taux occ. machine %",round(res.get("taux_m",0),2)),
             ("Pièces / Heure",round(res.get("pieces_heure",0),2)),
             ("Pièces / Jour",round(res.get("pieces_jour",0),2)),
             ("Repos / heure (min)",round(res.get("repos_par_heure_min",0),2)),
             ("CODE TEMPS",res.get("code_temps","—"))]
    for i,(k,v) in enumerate(results):
        bg=WHITE if i%2==0 else LIGHT
        bold_v=(k in("Temps cycle final (s)","CODE TEMPS","Repos / heure (min)"))
        val(ws,r2+1+i,1,k,bg=bg,align="left")
        c=val(ws,r2+1+i,2,v,bg=bg,bold=bold_v)
        if bold_v: c.font=Font(name="Arial",bold=True,color=ACCENT,size=11)

    ws2=wb.create_sheet("Données saisies"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:K1")
    hdr(ws2,1,1,"DONNÉES SAISIES — TABLEAU DES ÉTAPES",bg=DARK,bold=True,size=12,align="center")
    ws2.row_dimensions[1].height=28
    headers=["Machine","Étape","Début (s)","Durée (s)","Fin (s)","TM","TT","TTM","TR","TZ","TF"]
    widths=[10,28,12,12,12,6,6,6,6,6,6]
    for ci,(ch,cw) in enumerate(zip(headers,widths)):
        hdr(ws2,2,ci+1,ch,bg=ACCENT,size=10)
        ws2.column_dimensions[get_column_letter(ci+1)].width=cw
    TYPE_C={"TM":"FFE0B2","TT":"BBDEFB","TTM":"CFD8DC","TR":"F5F5F5","TZ":"EEEEEE"}
    for ri,row in edited_df.iterrows():
        r_xl=ri+3; bg=WHITE if ri%2==0 else LIGHT
        tm=bool(row.get("TM",False)); tt=bool(row.get("TT",False))
        ttm=bool(row.get("TTM",False)); tr=bool(row.get("TR",False)); tz=bool(row.get("TZ",False))
        if ttm: bg=TYPE_C["TTM"]
        elif tm: bg=TYPE_C["TM"]
        elif tt: bg=TYPE_C["TT"]
        elif tr: bg=TYPE_C["TR"]
        elif tz: bg=TYPE_C["TZ"]
        cells=[row.get("Sys",""),str(row.get("Etape","")),
               row.get("Debut",0),row.get("Duree",0),row.get("Fin",0),
               "✓" if tm else "","✓" if tt else "","✓" if ttm else "",
               "✓" if tr else "","✓" if tz else "",
               "✓" if bool(row.get("TF",False)) else ""]
        for ci,cv in enumerate(cells):
            val(ws2,r_xl,ci+1,cv,bg=bg,align="left" if ci==1 else "center")
    leg_r=len(edited_df)+4; ws2.row_dimensions[leg_r].height=8; leg_r+=1
    hdr(ws2,leg_r,1,"LÉGENDE",bg=DARK,size=9)
    for li,(code,color,desc) in enumerate([("TM","FFE0B2","Temps Manuel"),("TT","BBDEFB","Temps Machine"),
                                            ("TTM","CFD8DC","Temps Parallèle"),("TR","F5F5F5","Temps Repos"),
                                            ("TZ","EEEEEE","Temps Masqué")]):
        c1=ws2.cell(row=leg_r+1+li,column=1,value=code)
        c1.font=Font(name="Arial",bold=True,size=9); c1.fill=PatternFill("solid",fgColor=color)
        c1.alignment=Alignment(horizontal="center",vertical="center")
        c2=ws2.cell(row=leg_r+1+li,column=2,value=desc)
        c2.font=Font(name="Arial",size=9); c2.fill=PatternFill("solid",fgColor=color)
        c2.alignment=Alignment(horizontal="left",vertical="center")

    ws3=wb.create_sheet("Simogramme"); ws3.sheet_view.showGridLines=False
    ws3.merge_cells("A1:L1")
    hdr(ws3,1,1,"SIMOGRAMME",bg=DARK,bold=True,size=14,align="center")
    ws3.row_dimensions[1].height=32; ws3.column_dimensions["A"].width=18
    xl_img=XLImage(io.BytesIO(img_bytes)); xl_img.width=1100; xl_img.height=367
    ws3.add_image(xl_img,"A3")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ===================================================
# LOGIN
# ===================================================

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    st.markdown("## Connexion - Simogramme")
    _,col2,_=st.columns([1,2,1])
    with col2:
        user=st.text_input("Utilisateur"); pwd=st.text_input("Mot de passe",type="password")
        if st.button("Se connecter"):
            if user=="admin" and pwd=="1234":
                st.session_state["logged_in"]=True; st.rerun()
            else:
                st.error("Identifiants incorrects")
    st.stop()

# ===================================================
# SESSION STATE
# ===================================================

for k,v in [("machines",["M1"]),("show_history",False),("excel_bytes",None),
             ("fig_bytes",None),("pending_save",None),("editor_version",0),
             ("loaded_config",None),("chrono_etapes",[]),("show_chrono",False)]:
    if k not in st.session_state:
        st.session_state[k]=v

# ===================================================
# APPLY LOADED CONFIG
# ===================================================

if st.session_state["loaded_config"] is not None:
    cfg=st.session_state["loaded_config"]
    for key,field,default in [
        ("num_of","numero_of",""),("ref_piece","reference_piece",""),
        ("num_machine","numero_machine",""),("pdc","pdc",""),
        ("num_art","numero_article","")]:
        st.session_state[key]=str(cfg.get(field,"") or "")
    for key,field,default in [
        ("coef_th","coef_temps_humain",1.0),
        ("coef_tc","coef_temps_cycle",1.0),
        ("heures","heures_travail",7.0)]:
        st.session_state[key]=float(cfg.get(field,default) or default)
    try:
        machines_r=json.loads(cfg.get("machines",'["M1"]'))
    except Exception:
        machines_r=["M1"]
    st.session_state["machines"]=machines_r
    try:
        donnees_str=cfg.get("donnees","")
        if donnees_str:
            df_all=pd.read_json(io.StringIO(donnees_str))
            rename_map={"TM":"TM 🕐","TT":"TT 🤖","TTM":"TTM ⚡","TR":"TR ☕","TZ":"TZ ⚫","TF":"TF 🎨"}
            df_all.rename(columns=rename_map,inplace=True)
            for m in machines_r:
                df_m=df_all[df_all["Sys"]==m].copy() if "Sys" in df_all.columns else df_all.copy()
                df_m.drop(columns=["Sys","Fin"],errors="ignore",inplace=True)
                df_m.reset_index(drop=True,inplace=True)
                st.session_state[f"init_data_{m}"]=df_m
    except Exception as e:
        st.warning(f"Impossible de restaurer les tables: {e}")
    st.session_state["editor_version"]+=1
    st.session_state["loaded_config"]=None

# ===================================================
# SIDEBAR
# ===================================================

with st.sidebar:
    st.image(LOGO_URL,width=220)
    st.title("Configuration")
    st.markdown("## Informations production")
    numero_of      =st.text_input("Numéro OF",           key="num_of")
    numero_article =st.text_input("Numéro d'article",    key="num_art")
    reference_piece=st.text_input("Référence pièce",     key="ref_piece")
    numéro_machine =st.text_input("Numéro de la machine",key="num_machine")
    pdc            =st.text_input("PDC",                 key="pdc")

    st.markdown("---")
    st.markdown("## Coefficients")

    st.markdown("### ⏱️ Coef. Temps Humain")
    st.caption("Appliqué sur le temps manuel opérateur (ex: 1.10 = +10%)")
    coef_temps_humain = st.number_input(
        "Coef. Temps Humain",
        min_value=0.01, max_value=10.0, value=1.0, step=0.05,
        key="coef_th",
        help="Remplace le coefficient JA. Multiplie le temps manuel de l'opérateur."
    )

    st.markdown("---")
    st.markdown("### 🔄 Coef. Temps Cycle")
    st.caption("Appliqué sur le temps cycle après coef. humain (ex: 1.50 = +50% → repos = 33% du nouveau TC)")
    coef_temps_cycle = st.number_input(
        "Coef. Temps Cycle",
        min_value=1.0, max_value=5.0, value=1.0, step=0.05,
        key="coef_tc",
        help="Remplace le coefficient REPO. Multiplie le temps cycle total."
    )

    if coef_temps_cycle > 1.0:
        pct_base  = round((coef_temps_cycle - 1) * 100, 1)
        pct_final = round((1 - 1/coef_temps_cycle) * 100, 1)
        st.info(f"➕ **+{pct_base}%** du TC de base ajouté\n\n⏸️ **{pct_final}%** du nouveau TC = repos")
    else:
        st.info("Coef = 1.0 → pas de temps de repos ajouté")

    st.markdown("---")
    heures_travail=st.number_input("Heures travail/jour",min_value=1.0,max_value=24.0,value=7.0,step=0.5,key="heures")

    st.markdown("---")
    if st.button("➕ Ajouter machine"):
        new_m=f"M{len(st.session_state['machines'])+1}"
        st.session_state["machines"].append(new_m); st.rerun()

    st.markdown("---")
    st.markdown("## Modules")
    if st.button("⏱️ Chronométrage"):
        st.session_state["show_chrono"]=not st.session_state.get("show_chrono",False)
        st.rerun()
    if st.button("📊 Voir historique"):
        st.session_state["show_history"]=True; st.rerun()
    if st.button("❌ Fermer historique"):
        st.session_state["show_history"]=False; st.rerun()
    st.caption(f"DB: {DB_PATH}")

# ===================================================
# HISTORIQUE
# ===================================================

if st.session_state["show_history"]:
    st.markdown("## Historique des simulations")
    df_hist=load_configurations()
    if not df_hist.empty:
        for _,row in df_hist.iterrows():
            art  =row.get("numero_article","") or "—"
            of_v =row.get("numero_of","") or "—"
            pdc_v=row.get("pdc","") or "—"
            m_v  =row.get("numero_machine","") or "—"
            ref_v=row.get("reference_piece","") or "—"
            d_v  =str(row.get("date",""))[:16]
            th_v =row.get("coef_temps_humain",1)
            tc_v =row.get("coef_temps_cycle",1)
            try:
                res=json.loads(row.get("resultats","{}"))
                tc_s=round(res.get("temps_cycle_final",0),2)
                ph  =round(res.get("pieces_heure",0),1)
                pj  =round(res.get("pieces_jour",0),1)
                code=res.get("code_temps","—")
                rph =round(res.get("repos_par_heure_min",0),1)
            except Exception:
                tc_s=ph=pj=rph=0; code="—"

            st.markdown(f"""
            <div class="sim-card">
              <div class="sim-article">{art}</div>
              <div class="sim-date">📅 {d_v}</div>
              <div class="sim-row">
                <span class="sim-badge"><b>OF</b> {of_v}</span>
                <span class="sim-badge"><b>PDC</b> {pdc_v}</span>
                <span class="sim-badge"><b>Machine</b> {m_v}</span>
                <span class="sim-badge"><b>Réf.</b> {ref_v}</span>
                <span class="sim-badge"><b>Coef TH</b> {th_v}</span>
                <span class="sim-badge"><b>Coef TC</b> {tc_v}</span>
              </div>
              <div class="sim-kpi">
                <div class="sim-kpi-item"><div class="sim-kpi-val">{tc_s} s</div><div class="sim-kpi-lbl">Temps cycle</div></div>
                <div class="sim-kpi-item"><div class="sim-kpi-val">{ph}</div><div class="sim-kpi-lbl">Pièces/h</div></div>
                <div class="sim-kpi-item"><div class="sim-kpi-val">{pj}</div><div class="sim-kpi-lbl">Pièces/jour</div></div>
                <div class="sim-kpi-item"><div class="sim-kpi-val" style="color:#f97316">{rph} min</div><div class="sim-kpi-lbl">Repos/heure</div></div>
                <div class="sim-kpi-item"><div class="sim-kpi-val" style="color:#f97316">{code}</div><div class="sim-kpi-lbl">Code temps</div></div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            col_load, col_del = st.columns([1,1])
            with col_load:
                if st.button("📂 Charger",key=f"load_{row['id']}",use_container_width=True):
                    st.session_state["loaded_config"]=row.to_dict()
                    st.session_state["show_history"]=False; st.rerun()
            with col_del:
                if st.button("🗑️ Supprimer",key=f"del_{row['id']}",use_container_width=True):
                    delete_configuration(int(row["id"])); st.rerun()
            st.markdown("<div style='height:4px'></div>",unsafe_allow_html=True)
    else:
        st.info("Aucune simulation sauvegardée")
        st.caption(f"DB: {DB_PATH}")
    st.markdown("---")

# ===================================================
# MODULE CHRONOMÉTRAGE — avec coef par séquence + sélecteur machine
# ===================================================

if st.session_state.get("show_chrono", False):

    st.markdown("""
    <style>
    .chrono-title {
        background: linear-gradient(90deg,#1f2937,#374151);
        color:white; border-radius:10px; padding:14px 20px;
        font-weight:700; font-size:18px; margin-bottom:16px; letter-spacing:0.5px;
    }
    .chrono-table-wrap { overflow-x: auto; }
    table.chrono-tbl {
        border-collapse: collapse; width:100%;
        font-size: 12px; font-family: Arial, sans-serif;
    }
    table.chrono-tbl th {
        background:#1f2937; color:white; padding:6px 8px;
        text-align:center; border:1px solid #374151; white-space:nowrap;
    }
    table.chrono-tbl th.seq-col { background:#374151; }
    table.chrono-tbl th.fixed-col { background:#f97316; color:white; }
    table.chrono-tbl td {
        border:1px solid #d1d5db; padding:4px 6px;
        text-align:center; background:white;
    }
    table.chrono-tbl tr:nth-child(even) td { background:#f9fafb; }
    table.chrono-tbl td.seq-label {
        background:#fef3c7; font-weight:700; color:#92400e; text-align:left; white-space:nowrap;
    }
    table.chrono-tbl td.stat-moy { background:#d1fae5; font-weight:700; color:#065f46; }
    table.chrono-tbl td.stat-dp { background:#fef9c3; color:#713f12; }
    table.chrono-tbl td.stat-je { background:#ede9fe; color:#4c1d95; font-weight:700; }
    table.chrono-tbl td.stat-freq { background:#e0f2fe; color:#0c4a6e; font-weight:700; }
    table.chrono-tbl td.stat-coef { background:#fce7f3; color:#9d174d; font-weight:700; }
    table.chrono-tbl td.stat-final { background:#dcfce7; font-weight:700; color:#15803d; font-size:13px; }
    table.chrono-tbl td.mach-op  { background:#e0e7ff; color:#3730a3; font-weight:700; }
    table.chrono-tbl td.mach-m1  { background:#dbeafe; color:#1d4ed8; font-weight:700; }
    table.chrono-tbl td.mach-m2  { background:#dcfce7; color:#15803d; font-weight:700; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="chrono-title">⏱️ Feuille de relevés chronométrés</div>', unsafe_allow_html=True)

    TYPE_OPTIONS = ["TM", "TT", "TTM", "TR", "TZ", "TF"]
    # Machine options for chronométrage: dynamic from session machines + Opérateur
    MACH_OPTIONS = ["Opérateur"] + st.session_state["machines"]
    N_MAX_COL = 15
    DEFAULT_N_COL = 5

    if "chrono_n_col" not in st.session_state:
        st.session_state["chrono_n_col"] = DEFAULT_N_COL
    if "chrono_etapes" not in st.session_state or not st.session_state["chrono_etapes"]:
        st.session_state["chrono_etapes"] = [
            {"nom":"Séquence 1","type":"TM","freq":1,"coef":1.0,"machine":"Opérateur","prises":[0.0]*DEFAULT_N_COL}
        ]

    n_col = st.session_state["chrono_n_col"]

    tb1,tb2,tb3,tb4,tb5 = st.columns([2,2,2,2,2])
    with tb1:
        if st.button("➕ Séquence", key="chrono_add_row", use_container_width=True):
            idx_new = len(st.session_state["chrono_etapes"])+1
            st.session_state["chrono_etapes"].append(
                {"nom":f"Séquence {idx_new}","type":"TM","freq":1,"coef":1.0,"machine":"Opérateur","prises":[0.0]*n_col}
            )
            st.rerun()
    with tb2:
        if st.button("➖ Séquence", key="chrono_del_row", use_container_width=True):
            if len(st.session_state["chrono_etapes"])>1:
                st.session_state["chrono_etapes"].pop()
                st.rerun()
    with tb3:
        if st.button("➕ Colonne", key="chrono_add_col", use_container_width=True):
            if n_col < N_MAX_COL:
                st.session_state["chrono_n_col"] = n_col+1
                for e in st.session_state["chrono_etapes"]:
                    e["prises"].append(0.0)
                st.rerun()
    with tb4:
        if st.button("➖ Colonne", key="chrono_del_col", use_container_width=True):
            if n_col > 1:
                st.session_state["chrono_n_col"] = n_col-1
                for e in st.session_state["chrono_etapes"]:
                    e["prises"] = e["prises"][:n_col-1]
                st.rerun()
    with tb5:
        if st.button("🗑️ Réinitialiser", key="chrono_reset", use_container_width=True):
            st.session_state["chrono_etapes"] = [
                {"nom":"Séquence 1","type":"TM","freq":1,"coef":1.0,"machine":"Opérateur","prises":[0.0]*DEFAULT_N_COL}
            ]
            st.session_state["chrono_n_col"] = DEFAULT_N_COL
            st.rerun()

    st.markdown(f"**{n_col} mesures par séquence** — colonnes : {n_col} | séquences : {len(st.session_state['chrono_etapes'])}")
    st.markdown("""
    <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;padding:10px 14px;margin-bottom:12px;font-size:12px;color:#0369a1;">
    <b>🔢 Coef. séquence</b> — Multiplie la durée finale (après ÷ fréquence) avant transfert vers le tableau.<br>
    <b>🏭 Machine</b> — Attribue chaque séquence à Opérateur, M1 ou M2 (détermine la ligne du simogramme).
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    etapes_del = []
    for idx, etape in enumerate(st.session_state["chrono_etapes"]):
        # Pad / trim prises
        while len(etape["prises"]) < n_col:
            etape["prises"].append(0.0)
        etape["prises"] = etape["prises"][:n_col]
        # Ensure new keys exist in old records
        if "coef" not in etape: etape["coef"] = 1.0
        if "machine" not in etape: etape["machine"] = "Opérateur"

        vals = [v for v in etape["prises"] if v > 0]
        moy_raw = sum(vals)/len(vals) if vals else 0.0
        freq    = etape.get("freq", 1) or 1
        coef_seq = etape.get("coef", 1.0) or 1.0
        is_tf   = etape["type"] == "TF"
        moy_freq = round(moy_raw / freq, 4) if is_tf else round(moy_raw, 4)
        moy_final = round(moy_freq * coef_seq, 4)
        dp      = round((sum((v-moy_raw)**2 for v in vals)/len(vals))**0.5, 4) if len(vals)>1 else 0.0
        n_valid = len(vals)

        # Machine color for header
        mach_cur = etape.get("machine","Opérateur")
        if mach_cur == "M1":
            mach_color = "#dbeafe"; mach_text = "#1d4ed8"
        elif mach_cur == "M2":
            mach_color = "#dcfce7"; mach_text = "#15803d"
        else:
            mach_color = "#e0e7ff"; mach_text = "#3730a3"

        # Row header with machine badge
        st.markdown(
            f'<div style="background:{mach_color};border-radius:8px;padding:6px 12px;margin-bottom:4px;">'
            f'<span style="font-weight:700;color:{mach_text};font-size:13px;">#{idx+1} — {etape["nom"]}</span>'
            f'&nbsp;&nbsp;<span style="background:{mach_text};color:white;border-radius:4px;padding:1px 8px;font-size:11px;">{mach_cur}</span>'
            f'</div>',
            unsafe_allow_html=True
        )

        # Controls row: name | type | freq | coef | machine | delete
        hc1,hc2,hc3,hc4,hc5,hc6 = st.columns([2.5, 1.0, 1.0, 1.0, 1.2, 0.4])
        with hc1:
            nom = st.text_input("Séquence", value=etape["nom"],
                                key=f"cn_{idx}", label_visibility="collapsed",
                                placeholder=f"Séquence {idx+1}")
            st.session_state["chrono_etapes"][idx]["nom"] = nom
        with hc2:
            t_sel = st.selectbox("Type", TYPE_OPTIONS,
                                 index=TYPE_OPTIONS.index(etape["type"]) if etape["type"] in TYPE_OPTIONS else 0,
                                 key=f"ct_{idx}", label_visibility="collapsed")
            st.session_state["chrono_etapes"][idx]["type"] = t_sel
        with hc3:
            if t_sel == "TF":
                freq_in = st.number_input("Fréq.", min_value=1, max_value=1000,
                                          value=int(etape.get("freq",1)),
                                          key=f"cf_{idx}", label_visibility="collapsed")
                st.session_state["chrono_etapes"][idx]["freq"] = freq_in
                freq = freq_in
                moy_freq = round(moy_raw / freq, 4) if moy_raw > 0 else 0.0
            else:
                st.session_state["chrono_etapes"][idx]["freq"] = 1
                st.markdown("<div style='padding-top:8px;color:#9ca3af;font-size:11px'>fréq. N/A</div>",
                            unsafe_allow_html=True)
        with hc4:
            # Coef par séquence
            coef_in = st.number_input("Coef", min_value=0.01, max_value=10.0,
                                      value=float(etape.get("coef",1.0)),
                                      step=0.05, key=f"ccoef_{idx}",
                                      label_visibility="collapsed",
                                      help="Multiplie la durée finale de cette séquence avant transfert")
            st.session_state["chrono_etapes"][idx]["coef"] = coef_in
            coef_seq = coef_in
            moy_final = round(moy_freq * coef_seq, 4)
        with hc5:
            # Machine selector — rebuild MACH_OPTIONS dynamically
            mach_opts = ["Opérateur"] + st.session_state["machines"]
            cur_mach = etape.get("machine","Opérateur")
            if cur_mach not in mach_opts: cur_mach = "Opérateur"
            mach_sel = st.selectbox("Machine", mach_opts,
                                    index=mach_opts.index(cur_mach),
                                    key=f"cmach_{idx}", label_visibility="collapsed")
            st.session_state["chrono_etapes"][idx]["machine"] = mach_sel
        with hc6:
            if st.button("🗑️", key=f"cdel_{idx}"):
                etapes_del.append(idx)

        # Measurement inputs
        pcols = st.columns(n_col)
        new_prises = []
        for pi in range(n_col):
            with pcols[pi]:
                vp = st.number_input(f"T{pi+1}", min_value=0.0,
                                     value=float(etape["prises"][pi]),
                                     key=f"cp_{idx}_{pi}",
                                     label_visibility="visible", step=0.01)
                new_prises.append(vp)
        st.session_state["chrono_etapes"][idx]["prises"] = new_prises

        # Stats bar
        freq_disp = f"÷{freq}" if is_tf else "—"
        coef_disp = f"×{round(coef_seq,3)}" if coef_seq != 1.0 else "×1 (aucun)"
        st.markdown(
            f'<div class="chrono-stats" style="margin-bottom:8px;">'
            f'<b>N={n_valid}</b> &nbsp;|&nbsp; '
            f'Moy.brute: <b>{round(moy_raw,4)} s</b> &nbsp;|&nbsp; '
            f'Freq: <b>{freq_disp}</b> &nbsp;→&nbsp; {round(moy_freq,4)} s &nbsp;|&nbsp; '
            f'Coef: <b>{coef_disp}</b> &nbsp;|&nbsp; '
            f'<span style="color:#15803d;font-weight:800;font-size:14px">✅ Final: {moy_final} s</span> &nbsp;|&nbsp; '
            f'DP: {dp} &nbsp;|&nbsp; JE: <b>{t_sel}</b> &nbsp;|&nbsp; '
            f'<span style="font-weight:700;color:{mach_text}">🏭 {mach_sel}</span>'
            f'</div>', unsafe_allow_html=True)
        st.markdown("---")

    if etapes_del:
        for i in sorted(etapes_del, reverse=True):
            st.session_state["chrono_etapes"].pop(i)
        st.rerun()

    # ---- RÉCAPITULATIF TABLE ----
    st.markdown("### 📋 Récapitulatif — Feuille de relevés")
    recap_html = '<div class="chrono-table-wrap"><table class="chrono-tbl"><thead><tr>'
    recap_html += '<th class="seq-col">Séquence</th>'
    recap_html += '<th class="seq-col">Machine</th>'
    for pi in range(n_col):
        recap_html += f'<th>T{pi+1}</th>'
    recap_html += '<th class="fixed-col">Freq</th>'
    recap_html += '<th class="fixed-col">Moy.brute</th>'
    recap_html += '<th class="fixed-col">Coef</th>'
    recap_html += '<th class="fixed-col">Durée finale</th>'
    recap_html += '<th class="fixed-col">DP</th>'
    recap_html += '<th class="fixed-col">JE</th>'
    recap_html += '</tr></thead><tbody>'

    for etape in st.session_state["chrono_etapes"]:
        vals = [v for v in etape["prises"] if v > 0]
        moy_raw = sum(vals)/len(vals) if vals else 0.0
        freq    = etape.get("freq",1) or 1
        coef_seq = etape.get("coef",1.0) or 1.0
        is_tf   = etape["type"] == "TF"
        moy_freq = round(moy_raw/freq, 4) if is_tf else round(moy_raw, 4)
        moy_final = round(moy_freq * coef_seq, 4)
        dp      = round((sum((v-moy_raw)**2 for v in vals)/len(vals))**0.5, 4) if len(vals)>1 else 0.0
        freq_disp = str(freq) if is_tf else "—"
        mach_val = etape.get("machine","Opérateur")
        if mach_val == "M1": mach_cls = "mach-m1"
        elif mach_val == "M2": mach_cls = "mach-m2"
        else: mach_cls = "mach-op"

        recap_html += f'<tr><td class="seq-label">{etape["nom"]}</td>'
        recap_html += f'<td class="{mach_cls}">{mach_val}</td>'
        for pi in range(n_col):
            v = etape["prises"][pi] if pi < len(etape["prises"]) else 0.0
            cell_val = f"{v:.2f}" if v > 0 else ""
            recap_html += f'<td>{cell_val}</td>'
        recap_html += f'<td class="stat-freq">{freq_disp}</td>'
        recap_html += f'<td class="stat-moy">{round(moy_raw,4)}</td>'
        recap_html += f'<td class="stat-coef">{round(coef_seq,3)}</td>'
        recap_html += f'<td class="stat-final">{moy_final}</td>'
        recap_html += f'<td class="stat-dp">{dp}</td>'
        recap_html += f'<td class="stat-je">{etape["type"]}</td>'
        recap_html += '</tr>'

    recap_html += '</tbody></table></div>'
    st.markdown(recap_html, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # ---- TRANSFER BUTTON — routes to each machine table ----
    if st.button("🚀 Transférer vers les tableaux", use_container_width=True, key="chrono_inject"):
        # Group rows by machine destination
        rows_by_machine = {}
        for etape in st.session_state["chrono_etapes"]:
            dest = etape.get("machine","Opérateur")
            # "Opérateur" type sequences always go to M1 (main table)
            if dest == "Opérateur":
                dest_key = "M1"
            else:
                dest_key = dest
            # Ensure dest_key is in machines list
            if dest_key not in st.session_state["machines"]:
                st.session_state["machines"].append(dest_key)

            vals_nz = [v for v in etape["prises"] if v > 0]
            moy_raw = round(sum(vals_nz)/len(vals_nz), 4) if vals_nz else 0.0
            freq    = etape.get("freq",1) or 1
            coef_seq = etape.get("coef",1.0) or 1.0
            is_tf   = etape["type"] == "TF"
            moy_freq = round(moy_raw/freq, 4) if is_tf else moy_raw
            duree   = round(moy_freq * coef_seq, 4)
            t = etape["type"]

            if dest_key not in rows_by_machine:
                rows_by_machine[dest_key] = []
            rows_by_machine[dest_key].append({
                "Etape": etape["nom"],
                "Debut": 0.0,
                "Duree": duree,
                "TM 🕐": t == "TM",
                "TT 🤖": t == "TT",
                "TTM ⚡": t == "TTM",
                "TR ☕": t == "TR",
                "TZ ⚫": t == "TZ",
                "TF 🎨": t == "TF",
            })

        # Compute cumulative Debut for each machine and store
        for dest_key, rows in rows_by_machine.items():
            debut = 0.0
            for r in rows:
                r["Debut"] = round(debut, 4)
                debut += r["Duree"]
            st.session_state[f"init_data_{dest_key}"] = pd.DataFrame(rows)

        st.session_state["editor_version"] += 1
        st.session_state["show_chrono"] = False
        machines_filled = list(rows_by_machine.keys())
        st.success(f"✅ Données transférées vers : {', '.join(machines_filled)}")
        st.rerun()

    st.markdown("---")

# ===================================================
# LÉGENDE
# ===================================================

st.markdown("### Légende des types de temps")
leg_cols=st.columns(5)
for col,(color,code,tip,label) in zip(leg_cols,[
    ("#ff8c00","TM","Temps Manuel - Opérateur seul","Temps manuel"),
    ("#1f4fff","TT","Temps Technologique - Machine seule","Temps machine"),
    ("#111827","TTM","Opérateur + machine simultanément","Temps parallèle"),
    ("#9ca3af","TR","Temps de Repos","Temps repos"),
    ("#e5e7eb","TZ","Temps Masqué - non productif","Temps masqué"),
]):
    with col:
        st.markdown(
            f'<div style="text-align:center;"><span class="legend-color" style="background-color:{color};"></span>'
            f'<strong>{code}</strong><span class="info-icon" title="{tip}">?</span>'
            f'<br><small>{label}</small></div>',unsafe_allow_html=True)
st.markdown("---")

# ===================================================
# TABLES
# ===================================================

BOOL_COLS=["TM 🕐","TT 🤖","TTM ⚡","TR ☕","TZ ⚫","TF 🎨"]
ver=st.session_state["editor_version"]

def make_empty_df():
    return pd.DataFrame([{"Etape":"","Debut":0.0,"Duree":0.0,
        "TM 🕐":False,"TT 🤖":False,"TTM ⚡":False,
        "TR ☕":False,"TZ ⚫":False,"TF 🎨":False}])

dfs_calc=[]

for m in st.session_state["machines"]:
    ct,cd=st.columns([6,1])
    with ct: st.subheader(f"Tableau {m}")
    with cd:
        if m!="M1":
            if st.button("🗑️",key=f"del_machine_{m}"):
                st.session_state["machines"].remove(m)
                st.session_state.pop(f"init_data_{m}",None); st.rerun()
        else:
            st.write("")

    if f"init_data_{m}" in st.session_state:
        initial=st.session_state[f"init_data_{m}"].copy()
        for bc in BOOL_COLS:
            if bc not in initial.columns: initial[bc]=False
            initial[bc]=initial[bc].fillna(False).astype(bool)
        for nc in ["Debut","Duree"]:
            if nc in initial.columns:
                initial[nc]=pd.to_numeric(initial[nc],errors='coerce').fillna(0.0)
    else:
        initial=make_empty_df()

    if "Duree" in initial.columns:
        debut_vals=pd.to_numeric(initial["Debut"],errors='coerce').fillna(0)
        duree_vals=pd.to_numeric(initial["Duree"],errors='coerce').fillna(0)
        if debut_vals.sum()==0 and len(initial)>1:
            cumul=0.0
            for idx2 in range(len(initial)):
                initial.at[idx2,"Debut"]=cumul
                cumul+=float(duree_vals.iloc[idx2]) if idx2<len(duree_vals) else 0

    df_out=st.data_editor(
        initial, num_rows="dynamic", key=f"editor_{m}_v{ver}",
        use_container_width=True,
        column_config={
            "Etape":   st.column_config.TextColumn("Description étape",width="medium"),
            "Debut":   st.column_config.NumberColumn("Début (s)"),
            "Duree":   st.column_config.NumberColumn("Durée (s)"),
            "TM 🕐":  st.column_config.CheckboxColumn("TM"),
            "TT 🤖":  st.column_config.CheckboxColumn("TT"),
            "TTM ⚡": st.column_config.CheckboxColumn("TTM"),
            "TR ☕":  st.column_config.CheckboxColumn("TR"),
            "TZ ⚫":  st.column_config.CheckboxColumn("TZ"),
            "TF 🎨":  st.column_config.CheckboxColumn("TF"),
        }
    )

    df_c=df_out.copy()
    df_c.columns=[c.split(' ')[0] if ' ' in c else c for c in df_c.columns]
    df_c["Debut"]=pd.to_numeric(df_c["Debut"],errors='coerce').fillna(0)
    df_c["Duree"]=pd.to_numeric(df_c["Duree"],errors='coerce').fillna(0)
    df_c["Fin"]=df_c["Debut"]+df_c["Duree"]
    df_c["Sys"]=m
    dfs_calc.append(df_c)

edited_df=pd.concat(dfs_calc,ignore_index=True) if dfs_calc else pd.DataFrame()

# ===================================================
# GENERATE
# ===================================================

if st.button("🚀 Générer le simogramme",use_container_width=True):
    if edited_df.empty or edited_df["Duree"].sum()==0:
        st.error("Veuillez saisir des données dans au moins une table"); st.stop()

    fig,ax=plt.subplots(figsize=(18,6))
    fig.patch.set_facecolor('white'); ax.set_facecolor('white'); ax.set_frame_on(False)
    machines=st.session_state["machines"]
    y_pos={}; step=0.6; h=0.22; y_op=0
    for i,m in enumerate(machines):
        y_pos[m]=step*((i//2)+1) if i%2==0 else -step*((i//2)+1)

    max_x=0
    # Accumulators — global (all machines merged for operator)
    tm_total_global=0; man_total_global=0; par_total_global=0
    rep_total_global=0; msk_total_global=0
    # Per-machine accumulators for TT (machine time)
    machine_tt = {m: 0 for m in machines}
    machine_ttm = {m: 0 for m in machines}

    COLORS={"TM":"#ff8c00","TT":"#1f4fff","TTM":"#111827","TR":"#9ca3af","TZ":"#e5e7eb"}

    def do_hatch(ax,rect,x,y,w,ht,sp=0.2):
        i=0
        while i<w+ht:
            ln,=ax.plot([x+i,x+i-ht],[y,y+ht],color="black",lw=0.6,alpha=0.6)
            ln.set_clip_path(rect); i+=sp

    for _,row in edited_df.iterrows():
        op=str(row["Etape"]) if pd.notna(row["Etape"]) else ""
        s=float(row["Debut"]); t=float(row["Duree"]); e=s+t; sy=str(row["Sys"])
        tm=bool(row.get("TM",False)); tt=bool(row.get("TT",False))
        ttm=bool(row.get("TTM",False)); tr=bool(row.get("TR",False))
        tz=bool(row.get("TZ",False)); tf=bool(row.get("TF",False))
        if tz:
            msk_total_global+=t
            r=Rectangle((s,y_op),t,h,facecolor=COLORS["TZ"],edgecolor="black",alpha=0.4)
            ax.add_patch(r)
            if tf and t>0: do_hatch(ax,r,s,y_op,t,h)
            max_x=max(max_x,e); continue
        if tt and not ttm:
            machine_tt[sy] = machine_tt.get(sy,0) + t
            yp=y_pos.get(sy,0)
            r=Rectangle((s,yp),t,h,facecolor=COLORS["TT"],edgecolor="black")
            ax.add_patch(r)
            if tf and t>0: do_hatch(ax,r,s,yp,t,h)
            max_x=max(max_x,e)
        elif tm and not ttm:
            man_total_global+=t
            r=Rectangle((s,y_op),t,h,facecolor=COLORS["TM"],edgecolor="black")
            ax.add_patch(r)
            if tf and t>0: do_hatch(ax,r,s,y_op,t,h)
            max_x=max(max_x,e)
        elif ttm:
            machine_ttm[sy] = machine_ttm.get(sy,0) + t
            par_total_global+=t
            yp=y_pos.get(sy,0)
            r=Rectangle((s,y_op),t,yp-y_op,facecolor="#FFFFFF00",edgecolor="black")
            ax.add_patch(r)
            ax.plot([s,s+t],[y_op,yp],color="black",lw=1.5)
            if tf and t>0: do_hatch(ax,r,s,y_op,t,abs(yp-y_op))
            max_x=max(max_x,e)
        elif tr:
            rep_total_global+=t
            r=Rectangle((s,y_op),t,h,facecolor=COLORS["TR"],edgecolor="black",alpha=0.6)
            ax.add_patch(r)
            if tf and t>0: do_hatch(ax,r,s,y_op,t,h)
            max_x=max(max_x,e)
        if t>=0.5 and op not in ("","nan"):
            ax.text(s+t/2,y_op-0.18,op,ha="center",fontsize=9)

    for m,y in y_pos.items():
        ax.hlines(y,0,max_x,color="black",lw=1.5)
        ax.text(-1.5,y,m,ha="right",fontsize=14,fontweight="bold")
    ax.hlines(y_op,0,max_x,color="black",lw=2)
    ax.text(-1.5,y_op,"Opérateur",ha="right",fontsize=16,fontweight="bold")
    ax.legend(handles=[
        Patch(facecolor=COLORS["TM"],edgecolor='black',label='TM - Temps Manuel'),
        Patch(facecolor=COLORS["TT"],edgecolor='black',label='TT - Temps Machine'),
        Patch(facecolor=COLORS["TTM"],edgecolor='black',label='TTM - Temps Parallèle'),
        Patch(facecolor=COLORS["TR"],edgecolor='black',label='TR - Temps Repos'),
        Patch(facecolor=COLORS["TZ"],edgecolor='black',alpha=0.4,label='TZ - Temps Masqué'),
    ],loc='upper right',framealpha=0.9)
    ax.set_xlim(-2,max_x+2); ax.set_ylim(-1.5,1.5); ax.set_yticks([])
    ax.set_xlabel("Temps (secondes)",fontsize=12,fontweight="bold")
    ax.grid(axis="x",alpha=0.2,linestyle="--")
    plt.tight_layout()

    # ===================================================
    # CALCULS PAR MACHINE — cycle indépendant par machine
    # L'opérateur est partagé (même TM pour tous les cycles)
    # ===================================================
    hum = man_total_global + par_total_global + msk_total_global
    man_th = man_total_global * coef_temps_humain

    # Per-machine cycle calculations
    per_machine_results = {}
    for m in machines:
        tm_m = machine_tt.get(m, 0) + machine_ttm.get(m, 0)   # TT_m + TTM_m
        cyc_brut_m  = tm_m + man_total_global
        cyc_th_m    = tm_m + man_th
        cyc_fin_m   = cyc_th_m * coef_temps_cycle

        taux_h_m = (hum / cyc_brut_m * 100) if cyc_brut_m > 0 else 0
        taux_m_m = (tm_m / cyc_brut_m * 100) if cyc_brut_m > 0 else 0
        p_h_m    = 3600 / cyc_fin_m if cyc_fin_m > 0 else 0
        p_j_m    = p_h_m * heures_travail

        temps_repos_piece_m   = cyc_fin_m - cyc_th_m
        repos_heure_s_m       = temps_repos_piece_m * p_h_m
        repos_par_heure_min_m = repos_heure_s_m / 60
        pct_repos_m           = round((1 - 1/coef_temps_cycle)*100,1) if coef_temps_cycle > 1 else 0

        pe_m=int(cyc_fin_m); fr_m=cyc_fin_m-pe_m; m5_m=round(fr_m*20)/20
        if m5_m>=1.0: m5_m=0.95; pe_m+=1
        fc_m=int(m5_m*100)
        code_m=f"{pe_m}A{'01' if fc_m==0 else str(fc_m).zfill(2)}"

        per_machine_results[m] = {
            "total_machine_time": tm_m,
            "total_tt_only": machine_tt.get(m,0),
            "total_ttm": machine_ttm.get(m,0),
            "total_operator_manual": man_total_global,
            "total_operator_parallel": par_total_global,
            "total_masked_time": msk_total_global,
            "total_repos_time": rep_total_global,
            "cyc_brut": cyc_brut_m,
            "cyc_th": cyc_th_m,
            "temps_cycle_final": cyc_fin_m,
            "taux_h": taux_h_m,
            "taux_m": taux_m_m,
            "pieces_heure": p_h_m,
            "pieces_jour": p_j_m,
            "code_temps": code_m,
            "repos_par_heure_min": repos_par_heure_min_m,
            "pct_repos_du_cycle": pct_repos_m,
            "temps_repos_par_piece_s": temps_repos_piece_m,
        }

    # Use M1 (first machine) as primary resultats for saving/Excel
    first_m = machines[0]
    resultats_dict = per_machine_results[first_m].copy()
    # Also store all machines
    resultats_dict["per_machine"] = {m: per_machine_results[m] for m in machines}

    # ===================================================
    # KPIs — shared operator section + per-machine cycles
    # ===================================================
    st.markdown("## Indicateurs de performance")

    # --- Shared operator KPIs ---
    st.markdown("### 👷 Opérateur (commun à toutes les machines)")
    oc1,oc2,oc3,oc4 = st.columns(4)
    def kpi(col,v,label,delta="",style=""):
        css = style if style else "metric-card"
        val_css = "metric-value-m1" if "m1" in css else ("metric-value-m2" if "m2" in css else "metric-value")
        lbl_css = "metric-label-m1" if "m1" in css else ("metric-label-m2" if "m2" in css else "metric-label")
        dlt_css = "metric-delta-m1" if "m1" in css else ("metric-delta-m2" if "m2" in css else "metric-delta")
        with col:
            st.markdown(f'<div class="{css}"><div class="{val_css}">{v}</div>'
                        f'<div class="{lbl_css}">{label}</div>'
                        f'<div class="{dlt_css}">{delta}</div></div>',unsafe_allow_html=True)

    with oc1:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{round(man_total_global,2)} s</div>'
                    f'<div class="metric-label">Temps manuel TM brut</div>'
                    f'<div class="metric-delta">×{round(coef_temps_humain,2)} = {round(man_th,2)} s</div></div>',
                    unsafe_allow_html=True)
    with oc2:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{round(par_total_global,2)} s</div>'
                    f'<div class="metric-label">Temps parallèle TTM total</div>'
                    f'<div class="metric-delta">réparti par machine</div></div>',
                    unsafe_allow_html=True)
    with oc3:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{round(msk_total_global,2)} s</div>'
                    f'<div class="metric-label">Temps masqué TZ</div></div>',
                    unsafe_allow_html=True)
    with oc4:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{round(rep_total_global,2)} s</div>'
                    f'<div class="metric-label">Temps repos TR (saisi)</div></div>',
                    unsafe_allow_html=True)

    # --- Per-machine KPI blocks ---
    machine_colors = ["m1","m2","m1","m2"]  # cycle colors
    for mi, m in enumerate(machines):
        res_m = per_machine_results[m]
        color_key = machine_colors[mi % len(machine_colors)]
        card_cls = f"metric-card-{color_key}"

        if color_key == "m1":
            header_style = "cycle-header-m1"
        else:
            header_style = "cycle-header-m2"

        st.markdown(f'<div class="{header_style}">🏭 Cycle — {m}</div>', unsafe_allow_html=True)

        kc1,kc2,kc3,kc4,kc5,kc6 = st.columns(6)
        cols_m = [kc1,kc2,kc3,kc4,kc5,kc6]
        kpis_m = [
            (f"{round(res_m['temps_cycle_final'],2)} s", "Temps cycle final", f"×{coef_temps_cycle} TC"),
            (f"{round(res_m['temps_cycle_final']/36,3)} UM", "Temps cycle (UM)", ""),
            (f"{round(res_m['total_machine_time'],2)} s", "Temps machine TT+TTM",
             f"TT:{round(res_m['total_tt_only'],2)} TTM:{round(res_m['total_ttm'],2)}"),
            (f"{round(res_m['taux_h'],2)} %", "Taux occ. opérateur",
             f"hum={round(hum,2)} s"),
            (f"{round(res_m['taux_m'],2)} %", "Taux occ. machine", ""),
            (f"{round(res_m['pieces_heure'],2)}", "Pièces / Heure", ""),
        ]
        for col_m, (v_m,lbl_m,dlt_m) in zip(cols_m, kpis_m):
            with col_m:
                st.markdown(
                    f'<div class="{card_cls}"><div class="metric-value-{color_key}">{v_m}</div>'
                    f'<div class="metric-label-{color_key}">{lbl_m}</div>'
                    f'<div class="metric-delta-{color_key}">{dlt_m}</div></div>',
                    unsafe_allow_html=True)

        kd1,kd2,kd3 = st.columns(3)
        with kd1:
            st.markdown(
                f'<div class="{card_cls}"><div class="metric-value-{color_key}">{round(res_m["pieces_jour"],2)}</div>'
                f'<div class="metric-label-{color_key}">Pièces / Jour</div></div>',
                unsafe_allow_html=True)
        with kd2:
            st.markdown(
                f'<div class="{card_cls}"><div class="metric-value-{color_key}">{round(res_m["repos_par_heure_min"],2)} min</div>'
                f'<div class="metric-label-{color_key}">Repos / heure</div>'
                f'<div class="metric-delta-{color_key}">{res_m["pct_repos_du_cycle"]}% du cycle</div></div>',
                unsafe_allow_html=True)
        with kd3:
            st.markdown(
                f'<div class="{card_cls}"><div class="metric-value-{color_key}">{res_m["code_temps"]}</div>'
                f'<div class="metric-label-{color_key}">Code temps</div></div>',
                unsafe_allow_html=True)

        with st.expander(f"Détail des calculs — {m}"):
            for k_d,v_d in [
                ("TM (brut)",f"{round(man_total_global,4)} s"),
                ("TM × Coef TH",f"{round(man_th,2)} s"),
                (f"TT ({m})",f"{round(res_m['total_tt_only'],4)} s"),
                (f"TTM ({m})",f"{round(res_m['total_ttm'],4)} s"),
                (f"Temps machine total ({m})",f"{round(res_m['total_machine_time'],4)} s"),
                ("Cycle brut",f"{round(res_m['cyc_brut'],4)} s"),
                ("Cycle après Coef TH",f"{round(res_m['cyc_th'],2)} s"),
                (f"Coef Temps Cycle",f"{coef_temps_cycle:.4f}"),
                ("Cycle final",f"{round(res_m['temps_cycle_final'],2)} s"),
                ("Repos par pièce",f"{round(res_m['temps_repos_par_piece_s'],2)} s"),
                ("Pièces/heure",f"{round(res_m['pieces_heure'],4)}"),
                ("Repos total/heure",f"{round(res_m['repos_par_heure_min']*60,1)} s = {round(res_m['repos_par_heure_min'],2)} min"),
                ("CODE TEMPS",res_m["code_temps"])
            ]:
                st.write(f"**{k_d}:** {v_d}")

    st.success("✅ Simogramme généré avec succès")
    st.pyplot(fig)

    img_buf=io.BytesIO()
    fig.savefig(img_buf,format='png',bbox_inches="tight",dpi=150,facecolor='white')
    img_bytes=img_buf.getvalue()

    sidebar_info={
        "numero_of":numero_of,"reference_piece":reference_piece,
        "numero_machine":numéro_machine,"pdc":pdc,
        "numero_article":numero_article,
        "date":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "coef_temps_humain":coef_temps_humain,
        "coef_temps_cycle":coef_temps_cycle,
        "heures_travail":heures_travail,
    }
    excel_bytes=build_excel(edited_df,machines,sidebar_info,resultats_dict,img_bytes)

    st.session_state["fig_bytes"]=img_bytes
    st.session_state["excel_bytes"]=excel_bytes
    st.session_state["pending_save"]={
        'date':str(datetime.now()),'numero_of':numero_of,
        'reference_piece':reference_piece,'numero_machine':numéro_machine,
        'pdc':pdc,'numero_article':numero_article,
        'coef_temps_humain':coef_temps_humain,
        'coef_temps_cycle':coef_temps_cycle,
        'heures_travail':heures_travail,
        'machines':json.dumps(st.session_state["machines"]),
        'donnees':edited_df.to_json(),
        'resultats':json.dumps(resultats_dict, default=str),
    }

# ===================================================
# DOWNLOAD BUTTONS
# ===================================================

if st.session_state["excel_bytes"] or st.session_state["fig_bytes"]:
    st.markdown("---")
    cb1,cb2,cb3=st.columns(3)
    with cb1:
        if st.button("💾 Sauvegarder",key="save_btn",use_container_width=True):
            if st.session_state["pending_save"] and save_configuration(st.session_state["pending_save"]):
                st.success("✅ Sauvegardé !")
    with cb2:
        if st.session_state["excel_bytes"]:
            st.download_button("📥 Télécharger Excel",data=st.session_state["excel_bytes"],
                file_name=f"simogramme_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
    with cb3:
        if st.session_state["fig_bytes"]:
            st.download_button("🖼️ Télécharger PNG",data=st.session_state["fig_bytes"],
                file_name=f"simogramme_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                mime="image/png",use_container_width=True)
